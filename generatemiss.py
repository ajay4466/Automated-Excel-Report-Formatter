import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# Define paths 
input_folder = r"D:\Ajay Soni\PYthings\MIS\InputFiles"
output_folder = r"D:\Ajay Soni\PYthings\MIS\Output"
os.makedirs(output_folder, exist_ok=True)

# Process each CSV file
csv_files = glob.glob(os.path.join(input_folder, "*.csv"))

if not csv_files:
    print("❌ No CSV files found in input folder.")
else:
    for file_path in csv_files:
        try:
            df = pd.read_csv(file_path)

            # Start building output DataFrame
            temp_df = pd.DataFrame()
            temp_df['Sr No'] = ""  # fixed header spacing
            temp_df['Shop Name'] = df.get('storeTitle', "").astype(str).str.title()
            temp_df['Count'] = ""  # to be filled later
            temp_df['Elements'] = df.get('Element Name', "")
            temp_df['Product Name'] = df.get('Brand Name', "").astype(str).str.title()

            # Measurements
            temp_df['W in Inch'] = df.get('Width In Inch', 0).round(2)
            temp_df['H in Inch'] = df.get('Height In Inch', 0).round(2)
            temp_df['W in Ft'] = (temp_df['W in Inch'] / 12).round(2)
            temp_df['H in Ft'] = (temp_df['H in Inch'] / 12).round(2)
            temp_df['Quantity'] = df.get('Quantity', 1).round(2)
            temp_df['Total Sqft'] = (temp_df['W in Ft'] * temp_df['H in Ft'] * temp_df['Quantity']).round(2)

            # Other fields
            temp_df['Remark'] = df.get('Additional Information', "")
            temp_df['Recce Done By'] = df.get('agentFirstName', "").astype(str) + " " + df.get('agentLastName', "").astype(str)
            temp_df['Sales Person'] = df.get('Sales Person Name', "")
            temp_df['Recce Date'] = df.get('auditedOn', "")
            temp_df['Vendor Detail'] = ""
            temp_df['Execution Status'] = ""
            temp_df['Execution Date'] = ""
            temp_df['Location'] = df.get('storeLocation', "")
            temp_df['Location Link'] = ""
            temp_df['Lat'] = df.get('latitude', "")
            temp_df['Long'] = df.get('longitude', "")
            temp_df['Address'] = df.get('storeAddress', "")
            temp_df['Pincode'] = df.get('storePincode', "")
            temp_df['Contact Number'] = df.get('Shop Owner Contact Number', "")
            temp_df['Contact Person'] = df.get("Shop Owner's Name", "")

            # --- Fill-down logic (Ctrl+D) ---
            temp_df['Contact Number'] = temp_df['Contact Number'].replace("", None).fillna(method='ffill')
            temp_df['Contact Person'] = temp_df['Contact Person'].replace("", None).fillna(method='ffill')

            # --- Sr No and Count Logic ---
            temp_df['Count'] = temp_df.groupby('Shop Name').cumcount() + 1
            temp_df['Sr No'] = (temp_df['Shop Name'] != temp_df['Shop Name'].shift()).cumsum()
            temp_df['Sr No'] = temp_df['Sr No'].mask(temp_df['Shop Name'] == temp_df['Shop Name'].shift(), "")

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Recce Report"

            # Write data
            for r in dataframe_to_rows(temp_df, index=False, header=True):
                ws.append(r)

            # Header for reference
            header = [cell.value for cell in ws[1]]

            # Define styles
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Bold headers + center align
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Format each cell (all middle-center aligned)
            fixed_2_decimal_cols = ['W in Ft', 'H in Ft', 'Total Sqft']

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx, cell in enumerate(row):
                    col_name = header[col_idx]

                    # 0.00 format for numeric columns
                    if col_name in fixed_2_decimal_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

                    # Apply center-middle alignment for all cells
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Apply border to every cell
                    cell.border = thin_border

            # ===== Add Total Row =====
            last_row = ws.max_row + 1
            total_sqft_col = header.index('Total Sqft') + 1  # Excel col index
            total_sqft_letter = get_column_letter(total_sqft_col)

            # Count of Sr No in first column
            ws.cell(row=last_row, column=1).value = f"=COUNTA(A2:A{last_row-1})"
            ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

            # Merge TOTAL label from B to J
            ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=10)
            ws.cell(row=last_row, column=2).value = "TOTAL"
            ws.cell(row=last_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=last_row, column=2).font = Font(bold=True)

            # Sum of Total Sqft
            ws.cell(row=last_row, column=total_sqft_col).value = f"=SUM({total_sqft_letter}2:{total_sqft_letter}{last_row-1})"
            ws.cell(row=last_row, column=total_sqft_col).number_format = '0.00'
            ws.cell(row=last_row, column=total_sqft_col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=last_row, column=total_sqft_col).font = Font(bold=True)

            # Borders for total row
            for col in range(1, ws.max_column + 1):
                ws.cell(row=last_row, column=col).border = thin_border
                ws.cell(row=last_row, column=col).alignment = Alignment(horizontal='center', vertical='center')

            # Save Excel file
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_path = os.path.join(output_folder, f"{base_name}_output.xlsx")
            wb.save(output_path)
            print(f"✅ Saved: {output_path}")

        except Exception as e:
            print(f"❌ Error processing {file_path}: {e}")
